""" 
Write a program to invert the row and column of the cells in the spreadsheet. For example, the value at row 5, column 3 will be at row 3, column 5
(and vice versa). This should be done for all cells in the spreadsheet. For
example, the “before” and “after” spreadsheets would look something like

"""
import sys, os, openpyxl
from openpyxl.utils.exceptions import InvalidFileException

def invert_matrix(sheet):
  max_column = sheet.max_column
  max_row = sheet.max_row
  wb_ans = openpyxl.Workbook()
  sheet_ans = wb_ans['Sheet']

  for i in range(1, max_column + 1):
    for j in range(1, max_row + 1):
      sheet_ans.cell(row=i, column=j).value = sheet.cell(row=j, column=i).value
      
  wb_ans.save('3_ans.xlsx')



if __name__ == "__main__":

  if len(sys.argv) > 1:
    try:
      file_name = sys.argv[1]
      wb = openpyxl.load_workbook(file_name)
      sheet = wb[wb.sheetnames[0]]
      invert_matrix(sheet)
      wb.save(file_name)
    except InvalidFileException as Ex:
      print("Invalid File ", Ex)
    