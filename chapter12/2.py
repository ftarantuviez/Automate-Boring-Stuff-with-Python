"""
Create a program blankRowInserter.py that takes two integers and a filename
string as command line arguments. Let’s call the first integer N and the second integer M. Starting at row N, the program should insert M blank rows
into the spreadsheet. For example, when the program is run like this:

"""
import sys, os, openpyxl
from openpyxl.utils import get_column_letter

def blankRowInserter(start, cuantity, sheet):
	max_column = sheet.max_column
	min_column = sheet.min_column

	for i in range(start, start + cuantity):
		for j in range(min_column, max_column + 1):
			sheet.cell(row=i, column=j).value = ""
	
				

if __name__ == '__main__':
		
	if len(sys.argv) == 4:
		[start, cuantity, file_name] = sys.argv[1:]

		assert os.path.exists(file_name) and file_name[-5 : -1] == '.xls', \
			"No file in " + file_name

		wb = openpyxl.load_workbook(file_name)
		sheet = wb[wb.sheetnames[0]]
		blankRowInserter(int(start), int(cuantity), sheet)
		wb.save(file_name)

	else:
		print('BRO, no correct args, try again ;p')
